#!/usr/bin/env python3
"""
main.py  –  FastAPI backend for ZTE-AIS Gulf Solar BTS Dashboard 2025
Run:  uvicorn main:app --reload --port 8000
"""

from fastapi import FastAPI, Query, Request, Response, Depends, HTTPException, status
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse
from starlette.middleware.base import BaseHTTPMiddleware
from typing import Optional, List
import pandas as pd
import numpy as np
import os, json, secrets, hashlib, re
from datetime import datetime, timedelta

app = FastAPI(title="ZTE-AIS Gulf Solar BTS API", version="2.0")

# ══════════════════════════════════════════════════════════════════
# AUTH CONFIG
# ══════════════════════════════════════════════════════════════════
USER_FILE = os.path.join(os.path.dirname(__file__), "UserLogin.xlsx")
SESSION_TTL_HOURS = 8          # session expires after 8 hours
COOKIE_NAME = "bts_session"

# In-memory session store  { token: { "user": str, "role": str, "exp": datetime } }
_sessions: dict = {}

def _load_users() -> dict:
    """Load users from UserLogin.xlsx → { username_lower: { password, role } }"""
    try:
        df = pd.read_excel(USER_FILE, sheet_name="Sheet1", header=0)
        df.columns = [c.strip() for c in df.columns]
        users = {}
        for _, row in df.iterrows():
            u = str(row.get("User", "")).strip()
            p = str(row.get("Password", "")).strip()
            r = str(row.get("Role", "Member")).strip()
            if u and u.lower() not in ("nan", ""):
                users[u.lower()] = {"username": u, "password": p, "role": r}
        return users
    except Exception:
        return {}

def _get_session(request: Request) -> Optional[dict]:
    token = request.cookies.get(COOKIE_NAME)
    if not token:
        return None
    sess = _sessions.get(token)
    if not sess:
        return None
    if datetime.utcnow() > sess["exp"]:
        _sessions.pop(token, None)
        return None
    return sess

def require_auth(request: Request) -> dict:
    sess = _get_session(request)
    if not sess:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return sess

def require_admin(request: Request) -> dict:
    sess = require_auth(request)
    if sess.get("role", "").lower() != "admin":
        raise HTTPException(status_code=403, detail="Admin required")
    return sess

# ── Auth Middleware: protect all /api/* and / routes ─────────────
PUBLIC_PATHS = {"/api/login", "/api/logout", "/static/login.html",
                "/favicon.ico"}

class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        # Allow public paths and static assets (except dashboard root)
        if (path in PUBLIC_PATHS
                or path.startswith("/static/")
                or path == "/favicon.ico"):
            return await call_next(request)
        # All other routes require valid session
        sess = _get_session(request)
        if not sess:
            # API calls → 401 JSON
            if path.startswith("/api/"):
                return JSONResponse({"detail": "Not authenticated"}, status_code=401)
            # Page requests → redirect to login
            return RedirectResponse(url="/static/login.html", status_code=302)
        return await call_next(request)

app.add_middleware(AuthMiddleware)
app.mount("/static", StaticFiles(directory="static"), name="static")

# ══════════════════════════════════════════════════════════════════
# AUTH ROUTES
# ══════════════════════════════════════════════════════════════════
@app.get("/", include_in_schema=False)
def root():
    return FileResponse("static/index.html")

@app.get("/login", include_in_schema=False)
def login_page():
    return FileResponse("static/login.html")

@app.post("/api/login")
async def api_login(request: Request, response: Response):
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    username = str(body.get("username", "")).strip()
    password = str(body.get("password", "")).strip()

    if not username or not password:
        raise HTTPException(status_code=400, detail="Username and password required")

    users = _load_users()
    user_rec = users.get(username.lower())

    if not user_rec or user_rec["password"] != password:
        raise HTTPException(status_code=401, detail="Invalid username or password")

    # Create session token
    token = secrets.token_hex(32)
    _sessions[token] = {
        "user":  user_rec["username"],
        "role":  user_rec["role"],
        "exp":   datetime.utcnow() + timedelta(hours=SESSION_TTL_HOURS),
    }

    resp = JSONResponse({"ok": True, "user": user_rec["username"], "role": user_rec["role"]})
    resp.set_cookie(
        key=COOKIE_NAME, value=token,
        httponly=True, samesite="lax",
        max_age=SESSION_TTL_HOURS * 3600,
    )
    return resp

@app.post("/api/logout")
def api_logout(request: Request, response: Response):
    token = request.cookies.get(COOKIE_NAME)
    if token:
        _sessions.pop(token, None)
    resp = JSONResponse({"ok": True})
    resp.delete_cookie(COOKIE_NAME)
    return resp

@app.get("/api/me")
def api_me(request: Request):
    sess = _get_session(request)
    if not sess:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return {"user": sess["user"], "role": sess["role"]}

# ── Excel path: support both local and cloud (env var override) ──
_excel_name = os.environ.get(
    "EXCEL_FILE",
    "ZTE-AIS-Gulf Solar BTS 2025 Overall Progress.xlsx"
)
EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), _excel_name)
HEADER_ROW = 5   # 0-indexed → row 6 in Excel

# ══════════════════════════════════════════════════════════════════
# DATA LOADING  (hot-reload when Excel changes)
# ══════════════════════════════════════════════════════════════════
_df_cache: pd.DataFrame | None = None
_mtime: float = 0.0

def get_df() -> pd.DataFrame:
    global _df_cache, _mtime
    try:
        mt = os.path.getmtime(EXCEL_FILE)
    except FileNotFoundError:
        return pd.DataFrame()
    if _df_cache is None or mt != _mtime:
        _df_cache = _load()
        _mtime = mt
    return _df_cache

def find_col(df, name):
    name_l = name.strip().lower()
    for c in df.columns:
        if c.strip().lower() == name_l:
            return c
    for c in df.columns:
        if name_l in c.strip().lower():
            return c
    return None

def _load() -> pd.DataFrame:
    raw = pd.read_excel(EXCEL_FILE, sheet_name="DATA", header=HEADER_ROW)
    raw.columns = [str(c).strip() for c in raw.columns]
    raw = raw.dropna(how="all")

    WANTED = [
        # Identity
        "Site Code", "REF.", "Region", "Zone", "Site Status",
        "Site Name (EN)", "Province (EN)", "Tambon", "Amphur",
        "latitude", "longitude",
        # Site Config
        "Final Solar Subrack", "Group", "SCN", "Sub-SCN",
        "Tower Type", "Solution", "kW (Install)", "LOT",
        "Station Type", "Site Type", "Site Terrain",
        "Rollout Target", "Order Structure Lot",
        # Survey
        "Survey Status", "Survey Subcon", "Survey Daily Plan",
        "Survey Plan", "Survey Actual", "Survey Plan WK", "Survey Actual WK",
        # ETSS
        "ETSS Status", "ETSS Submit Date", "ETSS Approve Date",
        "ETSS Submit Month", "ETSS Submit SLA", "ETSS Review SLA", "Approved by",
        # TSSR
        "TSSR Report Status",
        "TSSR Report Submit in ZTE iEPMS Date",
        "TSSR Report Approve in ZTE iEPMS Date",
        "TSSR Report Submit to AIS (Offline) Actual",
        "TSSR Report Approve to AIS (Offline)",
        "TSSR Report Upload to Drive",
        # RFI / Config
        "RFI Grouping", "Structure Code 1", "Structure Code 2",
        "Propose Structure Type", "Foundation Propose", "RFI Status",
        "Issue Grouping",
        "Go No Go Status", "iEPMS PBOM Status", "BOM by site Status",
        "AIS Confirm Solution Date", "PBOM Confirm Actual", "BOM by site actual date",
        "TE Subcon", "SE Owner",
        # Milestone dates
        "Dispatch Plan Date", "Dispatch Actual Date",
        "MOS Plan Date", "MOS Actual Date",
        "Install Plan Date", "Install Actual Date",
        "On Service Plan Date", "On Service Actual Date",
        "PAT Report Plan Date", "PAT Report Actual Date",
        "MR Plan Date", "MR Actual Date",
        # Milestone WKs
        "Dispatch Plan WK", "Dispatch Actual WK",
        "MOS Plan WK", "MOS Actual WK",
        "Install Plan WK", "Install Actual WK",
        "On Service Plan WK", "On Service Actual WK",
        "PAT Report Plan WK", "PAT Report Actual WK",
        "MR Plan WK",
    ]

    selected = {}
    for want in WANTED:
        c = find_col(raw, want)
        if c:
            selected[want] = raw[c]

    df = pd.DataFrame(selected)

    DATE_COLS = [
        "Survey Plan", "Survey Actual", "Survey Daily Plan",
        "ETSS Submit Date", "ETSS Approve Date",
        "TSSR Report Submit in ZTE iEPMS Date",
        "TSSR Report Approve in ZTE iEPMS Date",
        "TSSR Report Submit to AIS (Offline) Actual",
        "TSSR Report Approve to AIS (Offline)",
        "TSSR Report Upload to Drive",
        "Dispatch Plan Date", "Dispatch Actual Date",
        "MOS Plan Date", "MOS Actual Date",
        "Install Plan Date", "Install Actual Date",
        "On Service Plan Date", "On Service Actual Date",
        "PAT Report Plan Date", "PAT Report Actual Date",
        "MR Plan Date", "MR Actual Date",
    ]
    for col in DATE_COLS:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    # Clean WK columns – replace '-' or NaN with None
    WK_COLS = [c for c in df.columns if "WK" in c]
    for col in WK_COLS:
        if col in df.columns:
            df[col] = df[col].replace("-", None).replace("", None)
            df[col] = df[col].where(df[col].notna(), None)

    # Milestone status
    MILESTONES = [
        ("Dispatch",  "Dispatch Plan Date",    "Dispatch Actual Date"),
        ("MOS",       "MOS Plan Date",         "MOS Actual Date"),
        ("Install",   "Install Plan Date",     "Install Actual Date"),
        ("OnService", "On Service Plan Date",  "On Service Actual Date"),
        ("PAT",       "PAT Report Plan Date",  "PAT Report Actual Date"),
        ("MR",        "MR Plan Date",          "MR Actual Date"),
    ]
    def ms_status(plan, actual):
        if pd.notna(actual): return "Done"
        if pd.notna(plan):   return "In Progress"
        return "Pending"

    for ms, p, a in MILESTONES:
        pv = df[p] if p in df.columns else pd.Series([None]*len(df))
        av = df[a] if a in df.columns else pd.Series([None]*len(df))
        df[f"{ms}_Status"] = [ms_status(x, y) for x, y in zip(pv, av)]

    # Numeric lat/lon
    for col in ["latitude", "longitude"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    return df

# ── helpers ───────────────────────────────────────────────────────
def clean(v):
    if v is None: return None
    if isinstance(v, float) and np.isnan(v): return None
    if isinstance(v, pd.Timestamp):
        return v.strftime("%d/%m/%Y") if not pd.isnull(v) else None
    if v is pd.NaT: return None
    try:
        if pd.isnull(v): return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, (np.integer,)): return int(v)
    if isinstance(v, (np.floating,)): return None if np.isnan(v) else float(v)
    if v == "-": return None
    return v

def df_to_records(df: pd.DataFrame) -> list:
    out = []
    for _, row in df.iterrows():
        out.append({c: clean(row[c]) for c in df.columns})
    return out

def pct(a, b):
    return round(a / b * 100, 1) if b else 0.0

def apply_filters(df: pd.DataFrame, params: dict) -> pd.DataFrame:
    for field, values in params.items():
        if not values or field not in df.columns:
            continue
        df = df[df[field].astype(str).isin([str(v) for v in values])]
    return df


# ══════════════════════════════════════════════════════════════════
# ROUTES
# ══════════════════════════════════════════════════════════════════

# ── meta / filter options ─────────────────────────────────────────
@app.get("/api/filter_options")
def api_filter_options(filters: str = Query(default="{}")):
    """Return available filter choices that are valid given current selections.
    Cascading: each field's options are scoped to the already-filtered dataset."""
    df = get_df()
    active = json.loads(filters)

    def uniq_in(df_sub, col):
        if col not in df_sub.columns: return []
        vals = df_sub[col].dropna().astype(str).unique().tolist()
        return sorted([v for v in vals if v not in ('', '-', 'nan', 'None')])

    FILTER_FIELDS = [
        ("Region",               "Region"),
        ("Zone",                 "Zone"),
        ("Province (EN)",        "Province (EN)"),
        ("Site Status",          "Site Status"),
        ("Survey Subcon",        "Survey Subcon"),
        ("TE Subcon",            "TE Subcon"),
        ("ETSS Status",          "ETSS Status"),
        ("TSSR Report Status",   "TSSR Report Status"),
        ("RFI Status",           "RFI Status"),
        ("Install_Status",       "Install_Status"),
        ("Dispatch Plan WK",     "Dispatch Plan WK"),
        ("SCN",                  "SCN"),
        ("Group",                "Group"),
        ("Rollout Target",       "Rollout Target"),
    ]

    result = {}
    for (label, col) in FILTER_FIELDS:
        # Apply all OTHER active filters except the field itself
        others = {k: v for k, v in active.items() if k != col}
        df_sub = apply_filters(df, others)
        result[label] = uniq_in(df_sub, col)

    # Count how many sites match ALL current filters
    df_filtered = apply_filters(df, active)
    result["__count__"] = len(df_filtered)
    return result


@app.get("/api/meta")
def api_meta():
    df = get_df()
    def uniq(col):
        if col not in df.columns: return []
        vals = df[col].dropna().astype(str).unique().tolist()
        vals = [v for v in vals if v not in ('', '-', 'nan', 'None')]
        return sorted(vals)

    return {
        "total": len(df),
        "as_of": datetime.today().strftime("%d %b %Y"),
        "filters": {
            "Region":               uniq("Region"),
            "Zone":                 uniq("Zone"),
            "Province (EN)":        uniq("Province (EN)"),
            "Site Status":          uniq("Site Status"),
            "Survey Status":        uniq("Survey Status"),
            "Survey Subcon":        uniq("Survey Subcon"),
            "Survey Plan WK":       uniq("Survey Plan WK"),
            "Survey Actual WK":     uniq("Survey Actual WK"),
            "Final Solar Subrack":  uniq("Final Solar Subrack"),
            "Group":                uniq("Group"),
            "SCN":                  uniq("SCN"),
            "Sub-SCN":              uniq("Sub-SCN"),
            "ETSS Status":          uniq("ETSS Status"),
            "ETSS Submit Month":    uniq("ETSS Submit Month"),
            "ETSS Submit SLA":      uniq("ETSS Submit SLA"),
            "ETSS Review SLA":      uniq("ETSS Review SLA"),
            "Approved by":          uniq("Approved by"),
            "TSSR Report Status":   uniq("TSSR Report Status"),
            "RFI Grouping":         uniq("RFI Grouping"),
            "RFI Status":           uniq("RFI Status"),
            "Propose Structure Type": uniq("Propose Structure Type"),
            "Foundation Propose":   uniq("Foundation Propose"),
            "Go No Go Status":      uniq("Go No Go Status"),
            "iEPMS PBOM Status":    uniq("iEPMS PBOM Status"),
            "BOM by site Status":   uniq("BOM by site Status"),
            "TE Subcon":            uniq("TE Subcon"),
            "Rollout Target":       uniq("Rollout Target"),
            "Dispatch Plan WK":     uniq("Dispatch Plan WK"),
            "Install Plan WK":      uniq("Install Plan WK"),
            "Install Actual WK":    uniq("Install Actual WK"),
            "Dispatch_Status":      ["Done","In Progress","Pending"],
            "MOS_Status":           ["Done","In Progress","Pending"],
            "Install_Status":       ["Done","In Progress","Pending"],
            "OnService_Status":     ["Done","In Progress","Pending"],
            "PAT_Status":           ["Done","In Progress","Pending"],
            "MR_Status":            ["Done","In Progress","Pending"],
        }
    }

# ── KPI summary ───────────────────────────────────────────────────
@app.get("/api/kpi")
def api_kpi(filters: str = Query(default="{}")):
    df = get_df()
    df = apply_filters(df, json.loads(filters))
    n = len(df)

    active  = int(df["Site Status"].isin(["Active","Active_BKK"]).sum()) if "Site Status" in df.columns else 0
    cancel  = int((df["Site Status"] == "Cancel").sum()) if "Site Status" in df.columns else 0
    rollout = int(df["Dispatch Plan Date"].notna().sum()) if "Dispatch Plan Date" in df.columns else 0

    survey_done = int((df["Survey Status"] == "Survey Done").sum()) if "Survey Status" in df.columns else 0

    # ETSS: Submitted = ETSS Submit Date is not null
    # ETSS: Approved  = ETSS Approve Date is not null
    # ETSS: Rejected  = ETSS Status contains "Reject"
    etss_submit = int(df["ETSS Submit Date"].notna().sum())  if "ETSS Submit Date"  in df.columns else 0
    etss_app    = int(df["ETSS Approve Date"].notna().sum()) if "ETSS Approve Date" in df.columns else 0
    etss_reject = int(df["ETSS Status"].astype(str).str.lower().str.contains("reject").sum()) if "ETSS Status" in df.columns else 0
    # ETSS Pending = Survey Done - ETSS Submitted (floor 0)
    etss_pending = max(0, survey_done - etss_submit)

    # TSSR: Submitted = has Submit date in iEPMS OR AIS Offline
    # TSSR: Approved  = has Approve date in iEPMS OR AIS Offline
    _tssr_sub_iepms = df["TSSR Report Submit in ZTE iEPMS Date"].notna()          if "TSSR Report Submit in ZTE iEPMS Date"         in df.columns else pd.Series(False, index=df.index)
    _tssr_sub_ais   = df["TSSR Report Submit to AIS (Offline) Actual"].notna()    if "TSSR Report Submit to AIS (Offline) Actual"   in df.columns else pd.Series(False, index=df.index)
    _tssr_app_iepms = df["TSSR Report Approve in ZTE iEPMS Date"].notna()         if "TSSR Report Approve in ZTE iEPMS Date"        in df.columns else pd.Series(False, index=df.index)
    _tssr_app_ais   = df["TSSR Report Approve to AIS (Offline)"].notna()          if "TSSR Report Approve to AIS (Offline)"         in df.columns else pd.Series(False, index=df.index)
    tssr_submit = int((_tssr_sub_iepms | _tssr_sub_ais).sum())
    tssr_app    = int((_tssr_app_iepms | _tssr_app_ais).sum())
    # TSSR Pending Review = count rows where TSSR Report Status == "Pending Approve"
    tssr_pending = int((df["TSSR Report Status"].astype(str).str.strip() == "Pending Approve").sum()) if "TSSR Report Status" in df.columns else 0
    # TSSR Report Submit to AIS (Offline) Actual
    tssr_submit_ais = int(_tssr_sub_ais.sum())
    # TSSR Report Approve to AIS (Offline)
    tssr_app_ais = int(_tssr_app_ais.sum())
    # TSSR Report Upload to Drive
    tssr_upload_drive = int(df["TSSR Report Upload to Drive"].notna().sum()) if "TSSR Report Upload to Drive" in df.columns else 0

    dispatch_done  = int((df["Dispatch_Status"] == "Done").sum())
    mos_done       = int((df["MOS_Status"] == "Done").sum())
    install_done   = int((df["Install_Status"] == "Done").sum())
    onservice_done = int((df["OnService_Status"] == "Done").sum())
    pat_done       = int((df["PAT_Status"] == "Done").sum())
    mr_done        = int((df["MR_Status"] == "Done").sum())

    rfi_go = int((df["RFI Status"] == "Go").sum()) if "RFI Status" in df.columns else 0

    return {
        "total": n, "active": active, "cancel": cancel, "rollout": rollout,
        "survey_done": survey_done,      "survey_pct":        pct(survey_done, n),    "survey_gap":        n - survey_done,
        "etss_submit": etss_submit,      "etss_submit_pct":   pct(etss_submit, n),    "etss_submit_gap":   n - etss_submit,
        "etss_app": etss_app,            "etss_pct":          pct(etss_app, n),        "etss_gap":          n - etss_app,
        "etss_reject": etss_reject,      "etss_reject_pct":   pct(etss_reject, n),
        "etss_pending": etss_pending,
        "tssr_submit": tssr_submit,      "tssr_submit_pct":   pct(tssr_submit, n),    "tssr_submit_gap":   n - tssr_submit,
        "tssr_app": tssr_app,            "tssr_pct":          pct(tssr_app, n),        "tssr_gap":          n - tssr_app,
        "tssr_pending": tssr_pending,
        "tssr_submit_ais": tssr_submit_ais, "tssr_submit_ais_pct": pct(tssr_submit_ais, n),
        "tssr_app_ais": tssr_app_ais,       "tssr_app_ais_pct":    pct(tssr_app_ais, n),
        "tssr_upload_drive": tssr_upload_drive, "tssr_upload_drive_pct": pct(tssr_upload_drive, n),
        "rfi_go": rfi_go,                "rfi_go_pct":        pct(rfi_go, n),
        "dispatch_done": dispatch_done,  "dispatch_pct":      pct(dispatch_done, n),  "dispatch_gap":      n - dispatch_done,
        "mos_done": mos_done,            "mos_pct":           pct(mos_done, n),        "mos_gap":           n - mos_done,
        "install_done": install_done,    "install_pct":       pct(install_done, n),   "install_gap":       n - install_done,
        "onservice_done": onservice_done,"onservice_pct":     pct(onservice_done, n), "onservice_gap":     n - onservice_done,
        "pat_done": pat_done,            "pat_pct":           pct(pat_done, n),        "pat_gap":           n - pat_done,
        "mr_done": mr_done,              "mr_pct":            pct(mr_done, n),         "mr_gap":            n - mr_done,
    }

# ── Tab 1: Overview ───────────────────────────────────────────────
@app.get("/api/overview")
def api_overview(filters: str = Query(default="{}")):
    df = get_df()
    df = apply_filters(df, json.loads(filters))

    def count_by(col):
        if col not in df.columns: return {}
        return df[col].value_counts().to_dict()

    # Region summary
    regions = sorted(df["Region"].dropna().unique().tolist()) if "Region" in df.columns else []
    region_summary = []
    for r in regions:
        sub = df[df["Region"] == r]; n = len(sub)
        # ETSS Submit = rows that have a date in ETSS Submit Date
        etss_submit_col = find_col(sub, "ETSS Submit Date")
        etss_submit_cnt = int(sub[etss_submit_col].notna().sum()) if etss_submit_col else 0
        # ETSS Reject = rows where ETSS Status contains "Reject"
        etss_rej_col = find_col(sub, "ETSS Status")
        etss_reject_cnt = int(sub[etss_rej_col].astype(str).str.lower().str.contains("reject").sum()) if etss_rej_col else 0
        region_summary.append({
            "region": r, "total": n,
            "survey_done":    int((sub["Survey Status"] == "Survey Done").sum()) if "Survey Status" in sub.columns else 0,
            "etss_submit":    etss_submit_cnt,
            "etss_approved":  int((sub["ETSS Status"] == "Approved").sum()) if "ETSS Status" in sub.columns else 0,
            "etss_reject":    etss_reject_cnt,
            "tssr_approved":  int((sub["TSSR Report Status"] == "TSSR Approved").sum()) if "TSSR Report Status" in sub.columns else 0,
            "rfi_go":         int((sub["RFI Status"] == "Go").sum()) if "RFI Status" in sub.columns else 0,
            "dispatch_done":  int((sub["Dispatch_Status"] == "Done").sum()),
            "mos_done":       int((sub["MOS_Status"] == "Done").sum()),
            "install_done":   int((sub["Install_Status"] == "Done").sum()),
            "onservice_done": int((sub["OnService_Status"] == "Done").sum()),
            "pat_done":       int((sub["PAT_Status"] == "Done").sum()),
            "mr_done":        int((sub["MR_Status"] == "Done").sum()),
        })

    ms_names = ["Dispatch","MOS","Install","OnService","PAT","MR"]
    milestone_chart = {
        ms: {
            "Done":        int((df[f"{ms}_Status"] == "Done").sum()),
            "In Progress": int((df[f"{ms}_Status"] == "In Progress").sum()),
            "Pending":     int((df[f"{ms}_Status"] == "Pending").sum()),
        } for ms in ms_names
    }

    # Province top 15
    prov = {}
    if "Province (EN)" in df.columns:
        prov = df["Province (EN)"].value_counts().head(15).to_dict()

    # Monthly survey plan trend
    monthly_survey = {}
    if "Survey Plan" in df.columns:
        valid = df[df["Survey Plan"].notna()].copy()
        valid["ym"] = valid["Survey Plan"].dt.strftime("%Y-%m")
        monthly_survey = dict(sorted(valid.groupby("ym").size().to_dict().items()))

    # ── Burndown helpers ──────────────────────────────────────────
    import re as _re

    def normalize_wk(s):
        """Normalise any WK string → 'YYW##' format (e.g. '26W01').
        Handles: '26W11', '26W8' → '26W08'; '2026-W11' → '26W11'."""
        s = str(s).strip()
        # Short format already: 26W11 or 26W8
        m = _re.match(r'^(\d{2})W(\d{1,2})$', s)
        if m:
            return f"{m.group(1)}W{m.group(2).zfill(2)}"
        # ISO format: 2026-W11
        m2 = _re.match(r'^(\d{4})-W(\d{1,2})$', s)
        if m2:
            yy = m2.group(1)[2:]   # last 2 digits of year
            return f"{yy}W{m2.group(2).zfill(2)}"
        return None

    def cumulative_by_wk_col(wk_col):
        """Cumulative count keyed by 'YYW##' from a WK string column."""
        if wk_col not in df.columns:
            return {}
        valid = df[df[wk_col].notna()].copy()
        valid['_wk'] = valid[wk_col].astype(str).apply(normalize_wk)
        valid = valid[valid['_wk'].notna() & (valid['_wk'] != '-')]
        if valid.empty:
            return {}
        weekly = valid.groupby('_wk').size().sort_index()
        cum, out = 0, {}
        for wk, cnt in weekly.items():
            cum += int(cnt)
            out[wk] = cum
        return out

    def cumulative_weekly_from_date(date_col):
        """Cumulative count keyed by 'YYW##' derived from a date column."""
        if date_col not in df.columns:
            return {}
        valid = df[df[date_col].notna()].copy()
        valid['_wk'] = valid[date_col].dt.strftime("%y") + "W" + valid[date_col].dt.strftime("%V").str.zfill(2)
        weekly = valid.groupby('_wk').size().sort_index()
        cum, out = 0, {}
        for wk, cnt in weekly.items():
            cum += int(cnt)
            out[wk] = cum
        return out

    survey_burndown = {
        "plan":   cumulative_by_wk_col("Survey Plan WK"),
        "actual": cumulative_by_wk_col("Survey Actual WK"),
    }

    onservice_burndown = {
        "plan":   cumulative_by_wk_col("On Service Plan WK"),
        "actual": cumulative_by_wk_col("On Service Actual WK"),
    }

    # SCN breakdown stacked by OnService status
    scn_onservice = {}
    if "SCN" in df.columns:
        scns = sorted(df["SCN"].dropna().unique().tolist())
        for scn_val in scns:
            sub = df[df["SCN"] == scn_val]
            scn_onservice[str(scn_val)] = {
                "Done":    int((sub["OnService_Status"] == "Done").sum()),
                "Pending": int((sub["OnService_Status"] != "Done").sum()),
            }

    # On Service overall done/pending for pie
    n_total = len(df)
    os_done    = int((df["OnService_Status"] == "Done").sum())
    os_pending = n_total - os_done

    return {
        "region_summary":      region_summary,
        "milestone_chart":     milestone_chart,
        "zone_count":          dict(sorted(count_by("Zone").items(), key=lambda x: -x[1])[:20]),
        "site_status":         count_by("Site Status"),
        "province_top":        dict(sorted(prov.items(), key=lambda x: -x[1])[:15]),
        "scn_breakdown":       dict(sorted(count_by("SCN").items(), key=lambda x: -x[1])),
        "scn_onservice":       scn_onservice,
        "onservice_pie":       {"Done": os_done, "Pending": os_pending},
        "group_breakdown":     count_by("Group"),
        "monthly_survey":      monthly_survey,
        "survey_burndown":     survey_burndown,
        "onservice_burndown":  onservice_burndown,
    }

# ── Tab 2: Survey / ETSS / TSSR ──────────────────────────────────
@app.get("/api/survey")
def api_survey(filters: str = Query(default="{}")):
    df = get_df()
    df = apply_filters(df, json.loads(filters))

    def count_by(col):
        if col not in df.columns: return {}
        return df[col].value_counts().to_dict()

    # Survey subcon table
    subcons = sorted(df["Survey Subcon"].dropna().unique().tolist()) if "Survey Subcon" in df.columns else []
    subcon_table = []
    for sc in subcons:
        sub = df[df["Survey Subcon"] == sc]; n = len(sub)
        done = int((sub["Survey Status"] == "Survey Done").sum()) if "Survey Status" in sub.columns else 0
        etss_submit = int(sub["ETSS Submit Date"].notna().sum()) if "ETSS Submit Date" in sub.columns else 0
        etss_reject = int(sub["ETSS Status"].astype(str).str.lower().str.contains("reject").sum()) if "ETSS Status" in sub.columns else 0
        etss = int((sub["ETSS Status"] == "Approved").sum()) if "ETSS Status" in sub.columns else 0
        _tssr_sub_iepms = sub["TSSR Report Submit in ZTE iEPMS Date"].notna() if "TSSR Report Submit in ZTE iEPMS Date" in sub.columns else pd.Series(False, index=sub.index)
        _tssr_sub_ais   = sub["TSSR Report Submit to AIS (Offline) Actual"].notna() if "TSSR Report Submit to AIS (Offline) Actual" in sub.columns else pd.Series(False, index=sub.index)
        tssr_submit = int((_tssr_sub_iepms | _tssr_sub_ais).sum())
        tssr = int((sub["TSSR Report Status"] == "TSSR Approved").sum()) if "TSSR Report Status" in sub.columns else 0
        days_list = []
        for _, row in sub.iterrows():
            p = row.get("Survey Plan"); a = row.get("Survey Actual")
            if pd.notna(p) and pd.notna(a):
                delta = (pd.Timestamp(a) - pd.Timestamp(p)).days
                days_list.append(delta)
        avg_days = round(sum(days_list)/len(days_list), 1) if days_list else None
        subcon_table.append({
            "subcon": sc, "total": n, "done": done, "pending": n - done,
            "pct": pct(done, n),
            "etss_submit": etss_submit, "etss_submit_pct": pct(etss_submit, n),
            "etss_pending": max(0, done - etss_submit),
            "etss_reject": etss_reject, "etss_reject_pct": pct(etss_reject, n),
            "etss": etss, "etss_pct": pct(etss, n),
            "tssr_submit": tssr_submit, "tssr_submit_pct": pct(tssr_submit, n),
            "tssr": tssr, "tssr_pct": pct(tssr, n),
        })

    # ETSS month (calendar-sorted)
    ORDER = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    etss_month_raw = count_by("ETSS Submit Month")
    etss_month = {m: etss_month_raw.get(m, 0) for m in ORDER if m in etss_month_raw}

    # Survey WK plan trend
    wk_plan = {}
    if "Survey Plan WK" in df.columns:
        wk_plan = df["Survey Plan WK"].dropna().value_counts().to_dict()
        wk_plan = dict(sorted(wk_plan.items()))

    # Survey by region
    regions = sorted(df["Region"].dropna().unique().tolist()) if "Region" in df.columns else []
    survey_region = []
    for r in regions:
        sub = df[df["Region"] == r]; n = len(sub)
        done         = int((sub["Survey Status"] == "Survey Done").sum()) if "Survey Status" in sub.columns else 0
        etss_submit  = int(sub["ETSS Submit Date"].notna().sum()) if "ETSS Submit Date" in sub.columns else 0
        etss_reject  = int(sub["ETSS Status"].astype(str).str.lower().str.contains("reject").sum()) if "ETSS Status" in sub.columns else 0
        etss_app     = int((sub["ETSS Status"] == "Approved").sum()) if "ETSS Status" in sub.columns else 0
        _tssr_sub_iepms = sub["TSSR Report Submit in ZTE iEPMS Date"].notna() if "TSSR Report Submit in ZTE iEPMS Date" in sub.columns else pd.Series(False, index=sub.index)
        _tssr_sub_ais   = sub["TSSR Report Submit to AIS (Offline) Actual"].notna() if "TSSR Report Submit to AIS (Offline) Actual" in sub.columns else pd.Series(False, index=sub.index)
        tssr_submit  = int((_tssr_sub_iepms | _tssr_sub_ais).sum())
        tssr_app     = int((sub["TSSR Report Status"] == "TSSR Approved").sum()) if "TSSR Report Status" in sub.columns else 0
        survey_region.append({
            "region": r, "total": n,
            "survey_done": done,         "survey_pct":       pct(done, n),
            "etss_submit": etss_submit,  "etss_submit_pct":  pct(etss_submit, n),
            "etss_pending": max(0, done - etss_submit),
            "etss_reject": etss_reject,  "etss_reject_pct":  pct(etss_reject, n),
            "etss_approved": etss_app,   "etss_pct":         pct(etss_app, n),
            "tssr_submit": tssr_submit,  "tssr_submit_pct":  pct(tssr_submit, n),
            "tssr_approved": tssr_app,   "tssr_pct":         pct(tssr_app, n),
        })

    # ── ETSS Submission SLA Pivot: Region × Survey Subcon × SLA category ──
    # Only show "pending" SLA rows (exclude already-approved, pending-survey, and submitted)
    EXCLUDE_SLA = {"ETSS Approved", "Pending Survey", "Submitted In SLA", "Submitted over SLA"}
    SLA_ORDER = [
        "Reject: Re-Submit Aging 1-3 Days",
        "Reject: Re-Submit Aging 4-6 Days",
        "Reject: Re-Submit Aging Over 6 Days",
        "Wait Submit ETSS Aging 1-3 Days",
        "Wait Submit ETSS Aging 4-6 Days",
        "Wait Submit ETSS Aging 7-9 Days",
        "Wait Submit ETSS Aging 10-12 Days",
        "Wait Submit ETSS Aging Over 12 Days",
    ]

    sla_subcons = sorted(df["Survey Subcon"].dropna().unique().tolist()) if "Survey Subcon" in df.columns else []
    sla_regions = ["BKK","CR","ER","NER","SR","NR"]
    sla_regions = [r for r in sla_regions if r in df["Region"].values] if "Region" in df.columns else []

    # Build pivot rows: per region, per SLA category
    etss_sla_pivot = []
    for region in sla_regions:
        reg_df = df[df["Region"] == region] if "Region" in df.columns else df
        # collect which SLA values exist in this region
        if "ETSS Submit SLA" not in reg_df.columns:
            continue
        sla_vals_in_region = set(reg_df["ETSS Submit SLA"].dropna().unique())
        active_slas = [s for s in SLA_ORDER if s in sla_vals_in_region and s not in EXCLUDE_SLA]
        if not active_slas:
            continue

        region_rows = []
        for sla in active_slas:
            sla_df = reg_df[reg_df["ETSS Submit SLA"] == sla]
            row = {"region": region, "sla": sla, "subcons": {}, "total": len(sla_df)}
            for sc in sla_subcons:
                cnt = int((sla_df["Survey Subcon"] == sc).sum()) if "Survey Subcon" in sla_df.columns else 0
                row["subcons"][sc] = cnt if cnt > 0 else None
            region_rows.append(row)

        # region subtotal
        reg_sla_df = reg_df[reg_df["ETSS Submit SLA"].isin(active_slas)] if "ETSS Submit SLA" in reg_df.columns else reg_df
        subtotal = {"region": region, "sla": "__total__", "subcons": {}, "total": len(reg_sla_df)}
        for sc in sla_subcons:
            cnt = int((reg_sla_df["Survey Subcon"] == sc).sum()) if "Survey Subcon" in reg_sla_df.columns else 0
            subtotal["subcons"][sc] = cnt if cnt > 0 else None
        region_rows.append(subtotal)
        etss_sla_pivot.extend(region_rows)

    # Grand total row
    all_pending_df = df[~df["ETSS Submit SLA"].isin(EXCLUDE_SLA)] if "ETSS Submit SLA" in df.columns else df
    grand_total = {"region": "__grand__", "sla": "__total__", "subcons": {}, "total": len(all_pending_df)}
    for sc in sla_subcons:
        cnt = int((all_pending_df["Survey Subcon"] == sc).sum()) if "Survey Subcon" in all_pending_df.columns else 0
        grand_total["subcons"][sc] = cnt if cnt > 0 else None
    etss_sla_pivot.append(grand_total)

    # ── Survey Burndown: remaining sites per week ──────────────────
    import re as _re

    def sort_wk_key(s):
        m = _re.match(r'^(\d+)W(\d+)$', str(s).strip())
        return (int(m.group(1)), int(m.group(2))) if m else (9999, 9999)

    def norm_wk(s):
        s = str(s).strip()
        m = _re.match(r'^(\d{2})W(\d{1,2})$', s)
        if m: return f"{m.group(1)}W{m.group(2).zfill(2)}"
        m2 = _re.match(r'^(\d{4})-W(\d{1,2})$', s)
        if m2: return f"{m2.group(1)[2:]}W{m2.group(2).zfill(2)}"
        return None

    # Sites that have a valid plan WK (i.e., assigned to survey schedule)
    valid_plan_mask  = df["Survey Plan WK"].notna() & (df["Survey Plan WK"].astype(str) != '-') if "Survey Plan WK" in df.columns else pd.Series(False, index=df.index)
    valid_actual_mask = df["Survey Actual WK"].notna() & (df["Survey Actual WK"].astype(str) != '-') if "Survey Actual WK" in df.columns else pd.Series(False, index=df.index)

    total_survey_planned = int(valid_plan_mask.sum())

    # Weekly counts per WK
    def weekly_counts(mask, wk_col):
        if wk_col not in df.columns: return {}
        sub = df[mask & df[wk_col].notna()].copy()
        sub['_wk'] = sub[wk_col].astype(str).apply(norm_wk)
        sub = sub[sub['_wk'].notna()]
        return sub.groupby('_wk').size().to_dict()

    wk_plan_counts   = weekly_counts(valid_plan_mask,   "Survey Plan WK")
    wk_actual_counts = weekly_counts(valid_actual_mask, "Survey Actual WK")

    all_wks = sorted(set(list(wk_plan_counts.keys()) + list(wk_actual_counts.keys())), key=sort_wk_key)

    survey_burndown_plan   = {}   # remaining plan per week
    survey_burndown_actual = {}   # remaining actual per week
    survey_weekly_plan     = {}   # weekly plan count (for bar)
    survey_weekly_actual   = {}   # weekly actual count (for bar)
    cum_p = 0
    cum_a = 0
    for wk in all_wks:
        wp = wk_plan_counts.get(wk, 0)
        wa = wk_actual_counts.get(wk, 0)
        cum_p += wp
        cum_a += wa
        survey_burndown_plan[wk]   = total_survey_planned - cum_p
        survey_burndown_actual[wk] = total_survey_planned - cum_a
        survey_weekly_plan[wk]     = int(wp)
        survey_weekly_actual[wk]   = int(wa)

    return {
        "survey_status":   count_by("Survey Status"),
        "etss_status":     count_by("ETSS Status"),
        "etss_month":      etss_month,
        "etss_submit_sla": count_by("ETSS Submit SLA"),
        "etss_review_sla": count_by("ETSS Review SLA"),
        "tssr_status":     count_by("TSSR Report Status"),
        "subcon_table":    subcon_table,
        "survey_region":   survey_region,
        "wk_plan":         wk_plan,
        "etss_sla_pivot":  etss_sla_pivot,
        "etss_sla_subcons": sla_subcons,
        "survey_burndown": {
            "total":         total_survey_planned,
            "weeks":         all_wks,
            "plan":          survey_burndown_plan,
            "actual":        survey_burndown_actual,
            "weekly_plan":   survey_weekly_plan,
            "weekly_actual": survey_weekly_actual,
        },
    }

# ── ETSS Submit Date Table ───────────────────────────────────────
@app.get("/api/etss_submit_date_table")
def api_etss_submit_date_table(
    filters: str = Query(default="{}"),
    months: str = Query(default="")      # comma-separated month labels, e.g. "Jan,Feb"
):
    df = get_df()
    df = apply_filters(df, json.loads(filters))

    MONTH_ORDER = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    # Available months from the data
    avail_months = []
    if "ETSS Submit Month" in df.columns:
        avail_months = [m for m in MONTH_ORDER if m in df["ETSS Submit Month"].dropna().unique()]

    # Selected months filter
    selected = [m.strip() for m in months.split(",") if m.strip()] if months else []
    if selected:
        fdf = df[df["ETSS Submit Month"].isin(selected)] if "ETSS Submit Month" in df.columns else df
    else:
        fdf = df  # no filter → show all

    # Get daily counts for each selected (or all) month-day
    # Columns: Region, Survey Subcon, then one column per date (or group by month if no filter)
    # We'll build: rows = region / subcon; cols = each ETSS Submit Date (date only)

    def safe_date(val):
        try:
            return pd.Timestamp(val).date()
        except Exception:
            return None

    # Build date list (sorted)
    if "ETSS Submit Date" in fdf.columns:
        dates_series = fdf["ETSS Submit Date"].dropna().apply(safe_date)
        dates_series = dates_series.dropna()
        all_dates = sorted(dates_series.unique())
    else:
        all_dates = []

    date_strs = [str(d) for d in all_dates]

    # Build rows grouped by Region + Survey Subcon
    rows = []
    # Get subcons and regions
    regions  = sorted(fdf["Region"].dropna().unique().tolist())       if "Region" in fdf.columns else []
    subcons  = sorted(fdf["Survey Subcon"].dropna().unique().tolist()) if "Survey Subcon" in fdf.columns else []

    # Grand totals per date
    grand = {"region": "Grand Total", "subcon": "", "pending": 0, "dates": {}}
    for d in date_strs:
        grand["dates"][d] = 0

    for region in regions:
        reg_df = fdf[fdf["Region"] == region]
        # region subtotal row
        region_row = {"region": region, "subcon": "", "pending": 0, "dates": {d: 0 for d in date_strs}}
        for sc in subcons:
            sub = reg_df[reg_df["Survey Subcon"] == sc] if "Survey Subcon" in reg_df.columns else reg_df
            if sub.empty:
                continue
            # pending = sites where survey done but ETSS not submitted
            survey_done_mask = (sub["Survey Status"] == "Survey Done") if "Survey Status" in sub.columns else pd.Series(False, index=sub.index)
            etss_not_sub = sub["ETSS Submit Date"].isna() if "ETSS Submit Date" in sub.columns else pd.Series(True, index=sub.index)
            pending = int((survey_done_mask & etss_not_sub).sum())

            # count by date
            date_counts = {}
            if "ETSS Submit Date" in sub.columns:
                valid = sub[sub["ETSS Submit Date"].notna()].copy()
                valid["_d"] = valid["ETSS Submit Date"].apply(safe_date).astype(str)
                date_counts = valid.groupby("_d").size().to_dict()

            sc_row = {"region": region, "subcon": sc, "pending": pending, "dates": {}}
            for d in date_strs:
                cnt = int(date_counts.get(d, 0))
                sc_row["dates"][d] = cnt if cnt else None
                region_row["dates"][d] += (cnt or 0)
                grand["dates"][d] += (cnt or 0)
            region_row["pending"] += pending
            grand["pending"] += pending
            rows.append(sc_row)

        # only append region row if has data
        if any(v > 0 for v in region_row["dates"].values()) or region_row["pending"] > 0:
            rows.insert(len(rows) - len([r for r in rows if r["region"] == region]), {**region_row, "_region_total": True})

    # Rebuild correctly: sort by region then subtotals first
    final_rows = []
    for region in regions:
        reg_rows = [r for r in rows if r["region"] == region]
        sub_rows = [r for r in reg_rows if not r.get("_region_total")]
        tot_rows = [r for r in reg_rows if r.get("_region_total")]
        if tot_rows or sub_rows:
            final_rows.extend(tot_rows)
            final_rows.extend(sub_rows)

    final_rows.append(grand)

    return {
        "avail_months": avail_months,
        "selected_months": selected if selected else avail_months,
        "dates": date_strs,
        "rows": final_rows,
    }


# ── Tab 3: Installation / Milestone ──────────────────────────────
@app.get("/api/installation")
def api_installation(filters: str = Query(default="{}")):
    df = get_df()
    df = apply_filters(df, json.loads(filters))

    ms_names  = ["Dispatch","MOS","Install","OnService","PAT","MR"]
    ms_labels = {"Dispatch":"Dispatch","MOS":"MOS","Install":"Install",
                 "OnService":"On Service","PAT":"PAT Report","MR":"MR"}

    milestone_stats = []
    for ms in ms_names:
        col = f"{ms}_Status"; total = len(df)
        done = int((df[col]=="Done").sum())
        ip   = int((df[col]=="In Progress").sum())
        pend = int((df[col]=="Pending").sum())
        milestone_stats.append({
            "name": ms_labels[ms], "key": ms,
            "done": done, "in_progress": ip, "pending": pend,
            "done_pct": pct(done, total),
        })

    # Milestone by Region
    regions = sorted(df["Region"].dropna().unique().tolist()) if "Region" in df.columns else []
    install_by_region = []
    for r in regions:
        sub = df[df["Region"]==r]; n = len(sub)
        install_by_region.append({
            "region": r, "total": n,
            "dispatch_done": int((sub["Dispatch_Status"]=="Done").sum()),
            "mos_done":      int((sub["MOS_Status"]=="Done").sum()),
            "install_done":  int((sub["Install_Status"]=="Done").sum()),
            "onservice_done":int((sub["OnService_Status"]=="Done").sum()),
            "pat_done":      int((sub["PAT_Status"]=="Done").sum()),
            "mr_done":       int((sub["MR_Status"]=="Done").sum()),
        })

    # Weekly dispatch plan
    weekly_dispatch_plan = {}
    if "Dispatch Plan WK" in df.columns:
        wk = df["Dispatch Plan WK"].dropna()
        weekly_dispatch_plan = dict(sorted(wk.value_counts().to_dict().items()))

    # Weekly install actual
    weekly_install_actual = {}
    if "Install Actual WK" in df.columns:
        wk2 = df["Install Actual WK"].dropna()
        weekly_install_actual = dict(sorted(wk2.value_counts().to_dict().items()))

    # Weekly install plan
    weekly_install_plan = {}
    if "Install Plan WK" in df.columns:
        wk3 = df["Install Plan WK"].dropna()
        weekly_install_plan = dict(sorted(wk3.value_counts().to_dict().items()))

    # Rollout target
    rollout_target = {}
    rollout_target_matrix  = {}   # { month: { region: count } }
    rollout_actual_matrix  = {}   # { month: { region: on_service_actual count } }
    rollout_months_order   = []
    rollout_regions_order  = []
    if "Rollout Target" in df.columns:
        rollout_target = df["Rollout Target"].dropna().value_counts().to_dict()
        if "Region" in df.columns:
            MONTH_ORDER = ["Jan","Feb","Mar","Apr","May","Jun",
                           "Jul","Aug","Sep","Oct","Nov","Dec"]
            _rt = df[df["Rollout Target"].notna() & df["Region"].notna()].copy()
            _rt["_month"]  = _rt["Rollout Target"].astype(str).str.strip()
            _rt["_region"] = _rt["Region"].astype(str).str.strip()
            for m in MONTH_ORDER:
                sub_m = _rt[_rt["_month"] == m]
                if sub_m.empty:
                    continue
                rollout_months_order.append(m)
                rollout_target_matrix[m] = sub_m.groupby("_region").size().to_dict()
                # on service actual = rows ที่มี On Service Actual Date
                if "On Service Actual Date" in df.columns:
                    act_m = sub_m[sub_m["On Service Actual Date"].notna()]
                    rollout_actual_matrix[m] = act_m.groupby("_region").size().to_dict()
                else:
                    rollout_actual_matrix[m] = {}
            rollout_regions_order = sorted(_rt["_region"].unique().tolist())

    # ── On Service Burndown (full, for Installation tab) ──────────────
    def norm_wk(s):
        s = str(s).strip().upper()
        m = re.match(r'^(\d{2})W(\d{1,2})$', s)
        if m: return f"{m.group(1)}W{int(m.group(2)):02d}"
        return None

    def sort_wk_key(s):
        m = re.match(r'^(\d{2})W(\d{2})$', s or '')
        return (int(m.group(1)), int(m.group(2))) if m else (99, 99)

    def weekly_os_counts(wk_col):
        if wk_col not in df.columns: return {}
        sub = df[df[wk_col].notna()].copy()
        sub['_wk'] = sub[wk_col].astype(str).apply(norm_wk)
        sub = sub[sub['_wk'].notna()]
        return sub.groupby('_wk').size().to_dict()

    valid_os_plan_mask = (
        df["On Service Plan WK"].notna() &
        (df["On Service Plan WK"].astype(str).str.strip() != '-')
    ) if "On Service Plan WK" in df.columns else pd.Series(False, index=df.index)

    total_os_planned = int(valid_os_plan_mask.sum())

    wk_os_plan_counts   = weekly_os_counts("On Service Plan WK")
    wk_os_actual_counts = weekly_os_counts("On Service Actual WK")

    all_os_wks = sorted(
        set(list(wk_os_plan_counts.keys()) + list(wk_os_actual_counts.keys())),
        key=sort_wk_key
    )

    os_burndown_plan = {}; os_burndown_actual = {}
    os_weekly_plan   = {}; os_weekly_actual   = {}
    cum_p = cum_a = 0
    for wk in all_os_wks:
        wp = wk_os_plan_counts.get(wk, 0)
        wa = wk_os_actual_counts.get(wk, 0)
        cum_p += wp; cum_a += wa
        os_burndown_plan[wk]   = total_os_planned - cum_p
        os_burndown_actual[wk] = total_os_planned - cum_a
        os_weekly_plan[wk]     = int(wp)
        os_weekly_actual[wk]   = int(wa)

    os_burndown = {
        "total":         total_os_planned,
        "weeks":         all_os_wks,
        "plan":          os_burndown_plan,
        "actual":        os_burndown_actual,
        "weekly_plan":   os_weekly_plan,
        "weekly_actual": os_weekly_actual,
    }

    return {
        "milestone_stats":       milestone_stats,
        "install_by_region":     install_by_region,
        "weekly_dispatch_plan":  weekly_dispatch_plan,
        "weekly_install_plan":   weekly_install_plan,
        "weekly_install_actual": weekly_install_actual,
        "rollout_target":        rollout_target,
        "rollout_target_matrix": rollout_target_matrix,
        "rollout_actual_matrix": rollout_actual_matrix,
        "rollout_months_order":  rollout_months_order,
        "rollout_regions_order": rollout_regions_order,
        "os_burndown":           os_burndown,
    }

# ══════════════════════════════════════════════════════════════════
# HLP PLAN  (stored as CSV on disk)
# ══════════════════════════════════════════════════════════════════
HLP_PLAN_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "HLP.csv")

REGION_ORDER_HLP = ["BKK", "CR", "ER", "NER", "NR", "SR"]

def _load_hlp_plan() -> dict:
    """Load HLP.csv → { region: { wk: int_or_null } }"""
    try:
        df = pd.read_csv(HLP_PLAN_FILE)
        df.columns = [str(c).strip() for c in df.columns]
        result = {}
        for _, row in df.iterrows():
            region = str(row.get("Region", "")).strip()
            if not region or region.lower() in ("nan", ""):
                continue
            wk_data = {}
            for col in df.columns:
                if col == "Region":
                    continue
                val = row[col]
                if pd.isna(val) or val == "" or val != val:
                    wk_data[col] = None
                else:
                    try:
                        wk_data[col] = int(val)
                    except (ValueError, TypeError):
                        wk_data[col] = None
            result[region] = wk_data
        return result
    except FileNotFoundError:
        return {}
    except Exception:
        return {}

def _save_hlp_plan(plan: dict):
    """Save plan dict → HLP.csv"""
    # Collect all week columns across regions
    all_wks = []
    for wk_data in plan.values():
        for wk in wk_data:
            if wk not in all_wks:
                all_wks.append(wk)
    # Sort weeks
    import re as _re
    def _wk_sort(s):
        m = _re.match(r'^(\d+)W(\d+)$', str(s).strip())
        return (int(m.group(1)), int(m.group(2))) if m else (9999, 9999)
    all_wks = sorted(all_wks, key=_wk_sort)
    rows = []
    for region in REGION_ORDER_HLP:
        if region not in plan:
            row = {"Region": region}
            for wk in all_wks:
                row[wk] = ""
            rows.append(row)
        else:
            row = {"Region": region}
            for wk in all_wks:
                v = plan[region].get(wk)
                row[wk] = int(v) if v is not None else ""
            rows.append(row)
    # Include extra regions not in REGION_ORDER_HLP
    for region in plan:
        if region not in REGION_ORDER_HLP:
            row = {"Region": region}
            for wk in all_wks:
                v = plan[region].get(wk)
                row[wk] = int(v) if v is not None else ""
            rows.append(row)
    out_df = pd.DataFrame(rows, columns=["Region"] + all_wks)
    out_df.to_csv(HLP_PLAN_FILE, index=False)

@app.get("/api/hlp/plan")
def api_hlp_plan_get(request: Request):
    """Get HLP weekly plan (all roles)."""
    require_auth(request)
    return _load_hlp_plan()

@app.post("/api/hlp/plan")
async def api_hlp_plan_save(request: Request):
    """Save HLP weekly plan (admin only)."""
    require_admin(request)
    body = await request.json()
    # body: { region: { wk: int_or_null } }
    _save_hlp_plan(body)
    return {"ok": True}

@app.get("/api/hlp/fill-from-plan-wk")
def api_hlp_fill_from_plan_wk(request: Request):
    """Compute weekly plan counts from 'On Service Plan WK' column grouped by Region (admin only)."""
    require_admin(request)
    df = get_df()

    def norm_wk(s):
        s = str(s).strip().upper()
        m = re.match(r'^(\d{2})W(\d{1,2})$', s)
        if m: return f"{m.group(1)}W{int(m.group(2)):02d}"
        return None

    # Find the relevant columns
    region_col = find_col(df, "Region")
    planwk_col = find_col(df, "On Service Plan WK")
    if not region_col or not planwk_col:
        raise HTTPException(status_code=400, detail="Required columns not found in DATA sheet")

    sub = df[[region_col, planwk_col]].dropna(subset=[planwk_col])
    result = {}  # { region: { wk: count } }
    for _, row in sub.iterrows():
        region = str(row[region_col]).strip()
        wk = norm_wk(str(row[planwk_col]))
        if not region or region.lower() in ('nan', '') or not wk:
            continue
        if region not in result:
            result[region] = {}
        result[region][wk] = result[region].get(wk, 0) + 1

    # Sort weeks within each region
    import re as _re
    def _wk_sort(s):
        m = _re.match(r'^(\d+)W(\d+)$', s)
        return (int(m.group(1)), int(m.group(2))) if m else (99, 99)
    for region in result:
        result[region] = dict(sorted(result[region].items(), key=lambda kv: _wk_sort(kv[0])))

    return result

@app.get("/api/hlp/plan/template")
def api_hlp_plan_template(request: Request):
    """Download blank HLP plan CSV template (admin only)."""
    require_admin(request)
    from fastapi.responses import StreamingResponse
    import io
    # Generate template with week columns from 26W01 to 26W53
    weeks = [f"26W{str(i).zfill(2)}" for i in range(1, 54)]
    rows = [{"Region": r, **{wk: "" for wk in weeks}} for r in REGION_ORDER_HLP]
    df_tmpl = pd.DataFrame(rows, columns=["Region"] + weeks)
    buf = io.StringIO()
    df_tmpl.to_csv(buf, index=False)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=HLP_Plan_Template.csv"}
    )

@app.post("/api/hlp/plan/import")
async def api_hlp_plan_import(request: Request):
    """Import HLP plan from CSV upload (admin only)."""
    from fastapi import UploadFile, File
    require_admin(request)
    body = await request.body()
    import io
    try:
        df = pd.read_csv(io.StringIO(body.decode("utf-8-sig")))
        df.columns = [str(c).strip() for c in df.columns]
        plan = {}
        for _, row in df.iterrows():
            region = str(row.get("Region", "")).strip()
            if not region or region.lower() in ("nan", ""):
                continue
            wk_data = {}
            for col in df.columns:
                if col == "Region":
                    continue
                val = row[col]
                if pd.isna(val) or val == "" or str(val).strip() == "":
                    wk_data[col] = None
                else:
                    try:
                        wk_data[col] = int(float(val))
                    except (ValueError, TypeError):
                        wk_data[col] = None
            plan[region] = wk_data
        _save_hlp_plan(plan)
        return {"ok": True, "regions": list(plan.keys())}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"CSV parse error: {e}")

# ── HLP Tab ──────────────────────────────────────────────────────
@app.get("/api/hlp")
def api_hlp(filters: str = Query(default="{}")):
    df = get_df()
    df = apply_filters(df, json.loads(filters))

    def norm_wk(s):
        s = str(s).strip().upper()
        m = re.match(r'^(\d{2})W(\d{1,2})$', s)
        if m: return f"{m.group(1)}W{int(m.group(2)):02d}"
        return None

    def sort_wk_key(s):
        m = re.match(r'^(\d{2})W(\d{2})$', s or '')
        return (int(m.group(1)), int(m.group(2))) if m else (99, 99)

    regions = sorted(df["Region"].dropna().unique().tolist()) if "Region" in df.columns else []

    # Get all On Service Actual WK values to determine week list
    os_actual_wks = set()
    if "On Service Actual WK" in df.columns:
        for v in df["On Service Actual WK"].dropna():
            w = norm_wk(str(v))
            if w: os_actual_wks.add(w)

    all_weeks = sorted(os_actual_wks, key=sort_wk_key)

    region_data = []
    for r in regions:
        sub = df[df["Region"] == r]
        scope = len(sub)
        actual_val = int((sub["OnService_Status"] == "Done").sum()) if "OnService_Status" in sub.columns else 0

        # weekly actual on service per region
        weekly_actual = {}
        if "On Service Actual WK" in sub.columns:
            for v in sub["On Service Actual WK"].dropna():
                w = norm_wk(str(v))
                if w:
                    weekly_actual[w] = weekly_actual.get(w, 0) + 1

        region_data.append({
            "region":        r,
            "scope":         scope,
            "actual":        actual_val,
            "weekly_actual": weekly_actual,
        })

    # Load HLP plan
    hlp_plan = _load_hlp_plan()

    # Merge HLP plan weeks into all_weeks
    plan_wks = set()
    for wk_data in hlp_plan.values():
        for wk in wk_data:
            w = norm_wk(wk)
            if w: plan_wks.add(w)
    all_weeks_raw = sorted(os_actual_wks | plan_wks, key=sort_wk_key)

    # Fill gaps: include every week between first and last week
    def fill_week_gaps(weeks):
        if len(weeks) < 2:
            return weeks
        import re as _re
        def wk_to_abs(s):
            m = _re.match(r'^(\d+)W(\d+)$', s)
            return int(m.group(1)) * 100 + int(m.group(2)) if m else None
        def abs_to_wk(n):
            yr, wn = divmod(n, 100)
            return f"{yr:02d}W{wn:02d}"
        nums = [wk_to_abs(w) for w in weeks if wk_to_abs(w) is not None]
        if not nums: return weeks
        full = []
        mn, mx = min(nums), max(nums)
        cur = mn
        while cur <= mx:
            yr, wn = divmod(cur, 100)
            # advance to next week: wn+1, handle year boundary (max 52/53 wks)
            full.append(abs_to_wk(cur))
            wn += 1
            if wn > 52:
                yr += 1; wn = 1
            cur = yr * 100 + wn
        return full

    all_weeks = fill_week_gaps(all_weeks_raw)

    # Attach weekly_plan to each region
    for rd in region_data:
        region = rd["region"]
        weekly_plan = {}
        if region in hlp_plan:
            for wk_raw, val in hlp_plan[region].items():
                wk = norm_wk(wk_raw)
                if wk and val is not None:
                    weekly_plan[wk] = val
        rd["weekly_plan"] = weekly_plan

    # Compute cumulative plan & actual per region
    for rd in region_data:
        cum_plan = {}; cum_actual = {}
        cum_p = 0; cum_a = 0
        for wk in all_weeks:
            cum_p += rd["weekly_plan"].get(wk, 0)
            cum_a += rd["weekly_actual"].get(wk, 0)
            if cum_p: cum_plan[wk] = cum_p
            if cum_a: cum_actual[wk] = cum_a
        rd["cum_plan"] = cum_plan
        rd["cum_actual"] = cum_actual

    # Grand total row
    total_scope = sum(rd["scope"] for rd in region_data)
    total_actual = sum(rd["actual"] for rd in region_data)
    grand_weekly_plan = {}; grand_weekly_actual = {}
    for rd in region_data:
        for wk, v in rd["weekly_plan"].items():
            grand_weekly_plan[wk] = grand_weekly_plan.get(wk, 0) + v
        for wk, v in rd["weekly_actual"].items():
            grand_weekly_actual[wk] = grand_weekly_actual.get(wk, 0) + v

    return {
        "weeks":        all_weeks,
        "regions":      region_data,
        "plan":         hlp_plan,
        "total_scope":  total_scope,
        "total_actual": total_actual,
        "grand_weekly_plan":   grand_weekly_plan,
        "grand_weekly_actual": grand_weekly_actual,
    }

# ── Tab 4: RFI + Config ───────────────────────────────────────────
@app.get("/api/rfi")
def api_rfi(filters: str = Query(default="{}")):
    df = get_df()
    df = apply_filters(df, json.loads(filters))

    def count_by(col):
        if col not in df.columns: return {}
        d = df[col].value_counts().to_dict()
        return {str(k): v for k, v in d.items() if str(k) not in ('-','nan','')}

    # RFI by region
    regions = sorted(df["Region"].dropna().unique().tolist()) if "Region" in df.columns else []
    rfi_region = []

    def col_count(sub, col, values=None, notnull=False):
        """Count rows where col matches any of values, or just is not-null/not-blank."""
        if col not in sub.columns: return 0
        if notnull:
            return int(sub[col].apply(lambda v: pd.notna(v) and str(v).strip() not in ('','-','nan','None','0')).sum())
        return int(sub[col].isin(values).sum())

    # Grand totals for later KPI use
    total_sites = len(df)
    gt_go         = col_count(df, "RFI Status",           ["Go"])
    gt_no_go      = col_count(df, "RFI Status",           ["No Go"])
    gt_wait       = col_count(df, "RFI Status",           ["Wait Review"])
    gt_etss_app   = col_count(df, "ETSS Status",          ["Approved"])
    gt_ais        = col_count(df, "AIS Confirm Solution Date", notnull=True)
    gt_pbom       = col_count(df, "PBOM Confirm Actual",        notnull=True)
    gt_pbom_pend  = max(0, total_sites - gt_go - gt_no_go)
    gt_bom_act    = col_count(df, "BOM by site actual date",    notnull=True)
    gt_bom_pend   = max(0, gt_pbom - gt_bom_act)

    for r in regions:
        sub = df[df["Region"]==r]; n = len(sub)
        go    = col_count(sub, "RFI Status",  ["Go"])
        no_go = col_count(sub, "RFI Status",  ["No Go"])
        wait  = col_count(sub, "RFI Status",  ["Wait Review"])
        known = {"Go", "No Go", "Wait Review"}
        if "RFI Status" in sub.columns:
            other = int(sub["RFI Status"].apply(
                lambda v: pd.notna(v) and str(v).strip() not in known and str(v).strip() not in ('', '-', 'nan', 'None')
            ).sum())
        else:
            other = 0
        etss_app  = col_count(sub, "ETSS Status",               ["Approved"])
        ais_conf  = col_count(sub, "AIS Confirm Solution Date",  notnull=True)
        pbom_conf = col_count(sub, "PBOM Confirm Actual",         notnull=True)
        pbom_pend = max(0, n - go - no_go)
        bom_act   = col_count(sub, "BOM by site actual date",     notnull=True)
        bom_pend  = max(0, pbom_conf - bom_act)
        rfi_region.append({
            "region": r, "total": n,
            "go": go, "no_go": no_go, "wait": wait, "other": other, "go_pct": pct(go, n),
            "etss_app": etss_app, "ais_conf": ais_conf,
            "pbom_conf": pbom_conf, "pbom_pend": pbom_pend,
            "bom_act": bom_act, "bom_pend": bom_pend,
        })

    # Issue Grouping by Region: { region: { issue_label: count } }
    issue_grouping_by_region = {}
    if "Issue Grouping" in df.columns and "Region" in df.columns:
        ig_df = df[df["Issue Grouping"].notna() & (df["Issue Grouping"].astype(str).str.strip().isin(['-','nan',''])==False)]
        for r in ig_df["Region"].dropna().unique():
            sub = ig_df[ig_df["Region"]==r]
            d = sub["Issue Grouping"].value_counts().to_dict()
            issue_grouping_by_region[str(r)] = {str(k): int(v) for k, v in d.items()}

    # Structure type detail
    struct_detail = []
    if "Propose Structure Type" in df.columns:
        for st, grp in df.groupby("Propose Structure Type"):
            if str(st) in ('-','nan',''): continue
            n = len(grp)
            go = int((grp["RFI Status"]=="Go").sum()) if "RFI Status" in grp.columns else 0
            struct_detail.append({
                "type": str(st), "total": n, "go": go, "go_pct": pct(go, n),
            })
        struct_detail.sort(key=lambda x: -x["total"])

    # ── ETSS Review SLA Pivot: Region × SE Owner × SLA category ──────
    REVIEW_SLA_COLS = [
        "Approved in 3 Days",
        "Approved in 4-6 Days",
        "Approved Over 6 Days",
        "Reviewing Aging 1-3 Days",
        "Reviewing Aging 4-6 Days",
        "Reviewing Aging Over 6 Days",
    ]
    REGION_ORDER = ["BKK", "CR", "ER", "NER", "SR", "NR"]

    etss_review_pivot = []
    etss_review_col_totals = {c: 0 for c in REVIEW_SLA_COLS}
    grand_total_count = 0

    if "ETSS Review SLA" in df.columns and "SE Owner" in df.columns and "Region" in df.columns:
        rev_df = df[df["ETSS Review SLA"].isin(REVIEW_SLA_COLS)].copy()
        rev_df["SE Owner"] = rev_df["SE Owner"].fillna("(blank)").astype(str)
        rev_df["Region"]   = rev_df["Region"].fillna("(blank)").astype(str)

        active_regions = [r for r in REGION_ORDER if r in rev_df["Region"].values]
        # also include any regions not in REGION_ORDER
        other_regions  = [r for r in sorted(rev_df["Region"].unique()) if r not in REGION_ORDER]
        all_regions    = active_regions + other_regions

        for region in all_regions:
            reg_df = rev_df[rev_df["Region"] == region]
            if reg_df.empty:
                continue
            se_owners = sorted(reg_df["SE Owner"].unique().tolist())

            region_subtotal = {c: 0 for c in REVIEW_SLA_COLS}
            detail_rows = []

            for se in se_owners:
                se_df = reg_df[reg_df["SE Owner"] == se]
                row_vals = {}
                row_total = 0
                for col in REVIEW_SLA_COLS:
                    cnt = int((se_df["ETSS Review SLA"] == col).sum())
                    row_vals[col] = cnt if cnt > 0 else None
                    region_subtotal[col] += cnt
                    row_total += cnt
                detail_rows.append({
                    "type":   "se",
                    "region": region,
                    "se":     se,
                    "vals":   row_vals,
                    "total":  row_total,
                })

            reg_total = sum(region_subtotal.values())
            sub_row = {
                "type":   "subtotal",
                "region": region,
                "se":     f"{region} Total",
                "vals":   {c: (v if v > 0 else None) for c, v in region_subtotal.items()},
                "total":  reg_total,
            }
            etss_review_pivot.extend(detail_rows)
            etss_review_pivot.append(sub_row)

            for c in REVIEW_SLA_COLS:
                etss_review_col_totals[c] += region_subtotal[c]
            grand_total_count += reg_total

    etss_review_pivot.append({
        "type":  "grand",
        "region": "__grand__",
        "se":    "Total",
        "vals":  {c: (v if v > 0 else None) for c, v in etss_review_col_totals.items()},
        "total": grand_total_count,
    })

    return {
        "total_sites":       len(df),
        "grand_totals": {
            "total": total_sites, "go": gt_go, "no_go": gt_no_go, "wait": gt_wait,
            "etss_app": gt_etss_app, "ais_conf": gt_ais,
            "pbom_conf": gt_pbom, "pbom_pend": gt_pbom_pend,
            "bom_act": gt_bom_act, "bom_pend": gt_bom_pend,
        },
        "rfi_status":        count_by("RFI Status"),
        "rfi_grouping":      count_by("RFI Grouping"),
        "issue_grouping":    count_by("Issue Grouping"),
        "issue_grouping_by_region": issue_grouping_by_region,
        "propose_structure": count_by("Propose Structure Type"),
        "foundation":        count_by("Foundation Propose"),
        "final_solar":       count_by("Final Solar Subrack"),
        "group_count":       count_by("Group"),
        "scn_count":         count_by("SCN"),
        "subscn_count":      count_by("Sub-SCN"),
        "bom_status":        count_by("BOM by site Status"),
        "go_no_go":          count_by("Go No Go Status"),
        "iepms_status":      count_by("iEPMS PBOM Status"),
        "te_subcon":         count_by("TE Subcon"),
        "order_lot":         count_by("Order Structure Lot"),
        "rfi_region":        rfi_region,
        "struct_detail":     struct_detail,
        "etss_review_pivot": etss_review_pivot,
        "etss_review_cols":  REVIEW_SLA_COLS,
    }

# ── Tab 5: Subcon performance ─────────────────────────────────────
@app.get("/api/subcon")
def api_subcon(filters: str = Query(default="{}")):
    df = get_df()
    df = apply_filters(df, json.loads(filters))

    subcons = sorted(df["Survey Subcon"].dropna().unique().tolist()) if "Survey Subcon" in df.columns else []
    rows = []
    for sc in subcons:
        sub = df[df["Survey Subcon"] == sc]; n = len(sub)
        done  = int((sub["Survey Status"] == "Survey Done").sum()) if "Survey Status" in sub.columns else 0
        etss  = int((sub["ETSS Status"] == "Approved").sum()) if "ETSS Status" in sub.columns else 0
        tssr  = int((sub["TSSR Report Status"] == "TSSR Approved").sum()) if "TSSR Report Status" in sub.columns else 0
        inst  = int((sub["Install_Status"] == "Done").sum())
        ons   = int((sub["OnService_Status"] == "Done").sum())
        rfi   = int((sub["RFI Status"] == "Go").sum()) if "RFI Status" in sub.columns else 0

        days_list = []
        for _, row in sub.iterrows():
            p = row.get("Survey Plan"); a = row.get("Survey Actual")
            if pd.notna(p) and pd.notna(a):
                days_list.append((pd.Timestamp(a) - pd.Timestamp(p)).days)
        avg_days = round(sum(days_list)/len(days_list), 1) if days_list else None

        region_breakdown = {}
        if "Region" in sub.columns:
            for r2 in sub["Region"].dropna().unique():
                rsub = sub[sub["Region"]==r2]
                rd = int((rsub["Survey Status"] == "Survey Done").sum()) if "Survey Status" in rsub.columns else 0
                region_breakdown[str(r2)] = {"total": len(rsub), "done": rd, "pct": pct(rd, len(rsub))}

        rows.append({
            "subcon": sc, "total": n,
            "survey_done": done, "survey_pending": n - done, "survey_pct": pct(done, n),
            "etss_approved": etss, "etss_pct": pct(etss, n),
            "tssr_approved": tssr, "tssr_pct": pct(tssr, n),
            "install_done": inst,  "install_pct": pct(inst, n),
            "onservice_done": ons, "onservice_pct": pct(ons, n),
            "rfi_go": rfi,         "rfi_go_pct": pct(rfi, n),
            "avg_survey_days": avg_days,
            "region_breakdown": region_breakdown,
        })

    chart = {
        "labels":        [r["subcon"] for r in rows],
        "survey_pct":    [r["survey_pct"] for r in rows],
        "etss_pct":      [r["etss_pct"] for r in rows],
        "tssr_pct":      [r["tssr_pct"] for r in rows],
        "install_pct":   [r["install_pct"] for r in rows],
        "onservice_pct": [r["onservice_pct"] for r in rows],
    }

    # TE Subcon table
    te_rows = []
    if "TE Subcon" in df.columns:
        for te, grp in df.groupby("TE Subcon"):
            if str(te) in ('-','nan',''): continue
            n = len(grp)
            inst = int((grp["Install_Status"]=="Done").sum())
            ons  = int((grp["OnService_Status"]=="Done").sum())
            te_rows.append({
                "te": str(te), "total": n,
                "install_done": inst, "install_pct": pct(inst, n),
                "onservice_done": ons, "onservice_pct": pct(ons, n),
            })
        te_rows.sort(key=lambda x: -x["total"])

    return {"table": rows, "chart": chart, "te_table": te_rows}

# ── Sites list (paginated) ────────────────────────────────────────
@app.get("/api/sites")
def api_sites(
    filters: str = Query(default="{}"),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=50, ge=1, le=500),
    search: str = Query(default=""),
    sort_by: str = Query(default="Site Code"),
    sort_dir: str = Query(default="asc"),
):
    df = get_df()
    df = apply_filters(df, json.loads(filters))
    if search:
        prov_mask = df["Province (EN)"].astype(str).str.lower().str.contains(search.lower(), na=False) if "Province (EN)" in df.columns else pd.Series([False]*len(df), index=df.index)
        code_mask = df["Site Code"].astype(str).str.lower().str.contains(search.lower(), na=False)
        df = df[code_mask | prov_mask]

    total = len(df)
    if sort_by in df.columns:
        df = df.sort_values(sort_by, ascending=(sort_dir=="asc"), na_position="last")

    start   = (page - 1) * page_size
    page_df = df.iloc[start:start+page_size]

    COLS = ["Site Code","Region","Zone","Province (EN)","Site Status",
            "Survey Status","ETSS Status","TSSR Report Status","RFI Status",
            "Install_Status","OnService_Status","PAT_Status","MR_Status",
            "latitude","longitude","Survey Subcon","TE Subcon"]
    COLS = [c for c in COLS if c in page_df.columns]

    return {"total": total, "page": page, "page_size": page_size,
            "data": df_to_records(page_df[COLS])}

# ── Single site detail ────────────────────────────────────────────
@app.get("/api/site/{site_code}")
def api_site_detail(site_code: str):
    df = get_df()
    rows = df[df["Site Code"].astype(str) == site_code]
    if rows.empty:
        return JSONResponse(status_code=404, content={"error": "Not found"})
    return df_to_records(rows)[0]

# ── Pivot Table ────────────────────────────────────────────────────
@app.get("/api/pivot")
def api_pivot(
    filters:    str = Query(default="{}"),
    row_fields: str = Query(default="Region"),      # comma-separated list
    col_fields: str = Query(default=""),            # comma-separated list
    val_type:   str = Query(default="count"),       # "count" | "pct"
):
    df = get_df()
    df = apply_filters(df, json.loads(filters))

    # Parse comma-separated field lists
    r_fields = [f.strip() for f in row_fields.split(",") if f.strip() and f.strip() in df.columns]
    c_fields = [f.strip() for f in col_fields.split(",") if f.strip() and f.strip() in df.columns]

    if not r_fields:
        return {"error": "No valid row_fields provided"}

    df = df.copy()
    # Fill NA for all grouping fields
    for f in r_fields + c_fields:
        df[f] = df[f].fillna("(blank)").astype(str)

    # Build column labels from col_fields (multi-level join)
    if c_fields:
        if len(c_fields) == 1:
            col_series = df[c_fields[0]]
        else:
            col_series = df[c_fields].apply(lambda r: " | ".join(r.values), axis=1)
        ct = pd.crosstab(
            [df[f] for f in r_fields],   # multi-level index
            col_series
        )
        col_labels = ct.columns.tolist()

        # Flatten multi-index → list of dicts with per-level keys
        rows_out = []
        for idx_tuple, row_data in ct.iterrows():
            if not isinstance(idx_tuple, tuple):
                idx_tuple = (idx_tuple,)
            counts = {c: int(row_data[c]) for c in col_labels}
            total  = sum(counts.values())
            values = counts
            if val_type == "pct" and total:
                values = {c: round(counts[c] / total * 100, 1) for c in col_labels}
            rows_out.append({
                "keys":   [str(k) for k in idx_tuple],   # one entry per r_field
                "values": values,
                "total":  total
            })

        # Sort: primary key desc-by-subtotal, then secondary keys desc-by-subtotal
        # Build subtotal map per level-0 key
        sub0 = {}
        for r in rows_out:
            sub0[r["keys"][0]] = sub0.get(r["keys"][0], 0) + r["total"]
        rows_out.sort(key=lambda r: (
            -sub0.get(r["keys"][0], 0),
            r["keys"][0],
            *[-sub for sub in [
                sum(x["total"] for x in rows_out
                    if x["keys"][:i+1] == r["keys"][:i+1])
                for i in range(1, len(r["keys"]))
            ]]
        ))

        totals = {c: int(ct[c].sum()) for c in col_labels}

    else:
        # Count only — group by all r_fields
        grp = df.groupby(r_fields).size().reset_index(name="__cnt__")
        col_labels = ["Count"]
        rows_out = []
        for _, row_data in grp.iterrows():
            keys = [str(row_data[f]) for f in r_fields]
            cnt  = int(row_data["__cnt__"])
            rows_out.append({"keys": keys, "values": {"Count": cnt}, "total": cnt})

        # Sort descending by level-0 subtotal, then level-1 subtotal, etc.
        sub0 = {}
        for r in rows_out:
            sub0[r["keys"][0]] = sub0.get(r["keys"][0], 0) + r["total"]
        rows_out.sort(key=lambda r: (
            -sub0.get(r["keys"][0], 0),
            r["keys"][0],
            *[-sub for sub in [
                sum(x["total"] for x in rows_out
                    if x["keys"][:i+1] == r["keys"][:i+1])
                for i in range(1, len(r["keys"]))
            ]]
        ))
        totals = {"Count": sum(r["total"] for r in rows_out)}

    return {
        "row_fields":  r_fields,
        "col_fields":  c_fields,
        "col_labels":  col_labels,
        "rows":        rows_out,   # each row has "keys" list instead of "row" string
        "grand_total": totals,
        "total_sites": len(df),
    }

# ── Map markers ───────────────────────────────────────────────────
@app.get("/api/map")
def api_map(filters: str = Query(default="{}")):
    df = get_df()
    df = apply_filters(df, json.loads(filters))
    MAP_COLS = ["Site Code","Site Name (EN)","Region","Zone","Province (EN)","Site Status",
                "Survey Status","ETSS Status","RFI Status",
                "Dispatch_Status","MOS_Status",
                "Install_Status","OnService_Status","PAT_Status","MR_Status",
                "Survey Subcon","TE Subcon",
                "latitude","longitude",
                "Survey Actual",
                "ETSS Approve Date",
                "Dispatch Actual Date","MOS Actual Date",
                "Install Actual Date","On Service Actual Date",
                "PAT Report Actual Date","MR Actual Date"]
    MAP_COLS = [c for c in MAP_COLS if c in df.columns]
    valid = df[df["latitude"].notna() & df["longitude"].notna()]
    return {"count": len(valid), "markers": df_to_records(valid[MAP_COLS])}

# ── Data Status (for auto-refresh polling) ────────────────────────
@app.get("/api/status")
def api_status():
    """Returns Excel file mtime + row count so frontend can detect updates."""
    try:
        mt    = os.path.getmtime(EXCEL_FILE)
        mtime_str = datetime.fromtimestamp(mt).strftime("%d/%m/%Y %H:%M:%S")
    except FileNotFoundError:
        mt, mtime_str = 0, "—"
    df = get_df()
    return {
        "mtime":      mt,
        "mtime_str":  mtime_str,
        "total_rows": len(df),
        "excel_file": os.path.basename(EXCEL_FILE),
    }

# ══════════════════════════════════════════════════════════════════
# ADMIN – User Management  (Admin role only)
# ══════════════════════════════════════════════════════════════════
import openpyxl

def _save_users(users: dict):
    """Persist users dict back to UserLogin.xlsx."""
    try:
        wb = openpyxl.load_workbook(USER_FILE)
        ws = wb["Sheet1"]
        # Clear existing rows (keep header row 1)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
        # Write users starting from row 2
        for i, (_, rec) in enumerate(users.items(), start=2):
            ws.cell(row=i, column=1, value=rec["username"])
            ws.cell(row=i, column=2, value=rec["password"])
            ws.cell(row=i, column=3, value=rec["role"])
        wb.save(USER_FILE)
        return True
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Save failed: {e}")

@app.get("/api/admin/users")
def admin_list_users(request: Request):
    """List all users (admin only)."""
    require_admin(request)
    users = _load_users()
    return [
        {"username": v["username"], "role": v["role"]}
        for v in users.values()
    ]

@app.post("/api/admin/users")
async def admin_create_user(request: Request):
    """Create a new user (admin only)."""
    require_admin(request)
    body = await request.json()
    username = str(body.get("username", "")).strip()
    password = str(body.get("password", "")).strip()
    role     = str(body.get("role", "Member")).strip()

    if not username or not password:
        raise HTTPException(status_code=400, detail="Username and password required")
    if role not in ("Admin", "Member"):
        raise HTTPException(status_code=400, detail="Role must be Admin or Member")

    users = _load_users()
    if username.lower() in users:
        raise HTTPException(status_code=409, detail="Username already exists")

    users[username.lower()] = {"username": username, "password": password, "role": role}
    _save_users(users)
    return {"ok": True, "username": username, "role": role}

@app.put("/api/admin/users/{username}")
async def admin_update_user(username: str, request: Request):
    """Update password and/or role (admin only)."""
    sess = require_admin(request)
    body = await request.json()
    new_password = str(body.get("password", "")).strip()
    new_role     = str(body.get("role", "")).strip()

    users = _load_users()
    key = username.lower()
    if key not in users:
        raise HTTPException(status_code=404, detail="User not found")

    # Prevent admin from demoting themselves
    if key == sess["user"].lower() and new_role and new_role != "Admin":
        raise HTTPException(status_code=400, detail="Cannot change your own role")

    if new_password:
        users[key]["password"] = new_password
    if new_role in ("Admin", "Member"):
        users[key]["role"] = new_role

    _save_users(users)
    return {"ok": True, "username": users[key]["username"], "role": users[key]["role"]}

@app.delete("/api/admin/users/{username}")
def admin_delete_user(username: str, request: Request):
    """Delete a user (admin only, cannot delete yourself)."""
    sess = require_admin(request)
    if username.lower() == sess["user"].lower():
        raise HTTPException(status_code=400, detail="Cannot delete your own account")

    users = _load_users()
    key = username.lower()
    if key not in users:
        raise HTTPException(status_code=404, detail="User not found")

    del users[key]
    _save_users(users)
    return {"ok": True}

@app.get("/admin", include_in_schema=False)
def admin_page():
    return FileResponse("static/admin.html")

# ── Export CSV ─────────────────────────────────────────────────────
@app.get("/api/export")
def api_export(filters: str = Query(default="{}")):
    from fastapi.responses import StreamingResponse
    import io
    df = get_df()
    df = apply_filters(df, json.loads(filters))

    EXPORT_COLS = ["Site Code","Region","Zone","Province (EN)","Site Status",
                   "Survey Status","Survey Subcon","ETSS Status","TSSR Report Status",
                   "RFI Status","Final Solar Subrack","Group","SCN","Sub-SCN",
                   "Dispatch_Status","MOS_Status","Install_Status",
                   "OnService_Status","PAT_Status","MR_Status",
                   "Install Plan Date","Install Actual Date",
                   "On Service Plan Date","On Service Actual Date"]
    cols = [c for c in EXPORT_COLS if c in df.columns]
    out = df[cols].copy()
    for col in out.select_dtypes(include="datetime64").columns:
        out[col] = out[col].dt.strftime("%d/%m/%Y")

    buf = io.StringIO()
    out.to_csv(buf, index=False)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=solar_bts_export.csv"}
    )
